import requests
import lxml.html
import sys
import codecs
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
import datetime


# input : UU:MM => YYYY/MM/DD UU:00:00
def convert_time_to_date_time(etime):
    etime_array = etime.split(":")
    #print(etime_array[0])
    #print(etime_array[1])
    etime = etime_array[0] + ":" + "00"
    #print(etime)
    etime += ':00'
    current_datetime = str(datetime.datetime.now())[0:10]
    current_datetime = current_datetime + " " + etime
    #print(current_datetime)
    return current_datetime

    #rcurrent_datetime.strftime('%x %X')

# zorgt ervoor dat speciale karakters geprint worden : vb. Chinese karakters
sys.stdout = codecs.getwriter('utf8')(sys.stdout.buffer)


html = requests.get('https://www.beursduivel.be/Koersen/Aandelen.aspx')
doc = lxml.html.fromstring(html.content)

Issue_Default = doc.xpath('//div[@class="IssueDefault"]')[0]
titles = Issue_Default.xpath('.//td[@class="TitleCell DateTimeCell"]/a/text()')
lastprices = Issue_Default.xpath('.//td[@class="ValueCell"][1]/span/text()')
highprices = Issue_Default.xpath('.//td[@class="ValueCell"][4]/span/text()')
lowprices = Issue_Default.xpath('.//td[@class="ValueCell"][5]/span/text()')
finals = Issue_Default.xpath('.//td[@class="ValueCell"][6]/span/text()')
times = Issue_Default.xpath('.//td[@class="ValueCell"][7]/span/text()')

#print(titles)
#print(prices)


output = []
for info in zip(titles, lastprices, highprices, lowprices,finals,times):
    resp = {}
    resp['title'] = info[0]
    resp['price'] = info[1]
    resp['high'] = info[2]
    resp['low'] = info[3]
    resp['final'] = info[4]
    resp['time'] = convert_time_to_date_time(info[5])
    output.append(resp)
#for var in enumerate(output):
#    print(var)

wb=load_workbook('d:\data\it\python\koersen.xlsx')

# activate demo.xlsx
sheet=wb.active

# get max row count
max_row=sheet.max_row

#workbook = xlsxwriter.Workbook('d:\data\it\python\koersen.xlsx')
#worksheet = workbook.add_worksheet()

workbook = openpyxl.load_workbook('d:\data\it\python\koersen.xlsx')
worksheet = workbook['Sheet1']

row = max_row

print(row)

#row = -1

for j in output:
        #print(j)
        row +=1
        col = 0
        for k in j:
            col += 1
            print(j[k])
            print(row)
            print(col)
            #print("{}:{}".format(k,j[k]))
            #worksheet.write(row,col,j[k] )
            worksheet.cell(row, col).value = j[k]

workbook.save('d:\data\it\python\koersen.xlsx')

workbook.close()
